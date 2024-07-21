import openpyxl as xls
from openpyxl.styles import PatternFill
from time import sleep
import datetime as time

#messages that will be presented/asked to the user
User_message = ''' 
ATENÇÃO: 
Para que este programa funcione:\n
não devem haver colunas mescladas na planilha original;
as datas de início e término de cada etapa devem estar numa mesma célula, no formato "dd/mm até dd/mm";
as datas devem estar arranjadas em colunas, não em linhas;
o cronograma montado irá apagar dados que estejam no espaço indicado pelo usuário.\n
Por favor, indique os seguintes dados (digite os dados pedidos, sem espaços no começo ou no final, e aperte Enter):\n'''
path_message = 'Caminho do arquivo (ex: C:/Usuários/USUARIO/Downloads/arquivo.xlsx) (OBS: use a barra assim / e não \): '
sheet_name_message = 'Nome da planilha (ex: Planilha1): '
planned_column_message = '\nColuna com as datas previstas no projeto (coloque um número, não a letra: coluna A é número 1, etc.): '
actual_column_message = 'Coluna com as datas realizadas na execução (coloque um número, não a letra: coluna A é número 1, etc.): '
first_colored_column_message = 'Coluna que indicará a primeira semana da obra (as outras estarão à direita)(coloque um número): '
top_line_message = '\nLinha que será o topo do cronograma, indicando as semanas (coloque um número): '
highlighted_lines_message = 'Linhas de destaque (terão cor diferente das outras) (coloque números separados por vírgula, sem espaços, ex: 5,17,23) (se não houverem, só aperte Enter): '
final_user_message = 'Pronto! Agora você já tem uma nova planilha atualizada no mesmo local que a original.'



year = time.datetime.now().year

# givens:
# planned_col, actual_col, top_line (contains the dates), file_path (from C:/), first_colored_col, sheet_name


def get_stage_dates(file_path, sheet_name, planned_col, actual_col, top_line):
    '''considera duas cells adjacentes com texto no formato "dd/mm até dd/mm" e assume que está no ano atual
    OBS: the actual begin and finish dates won't necessarily have a numerical value, they can be None. Just the planned ones are guaranteed to have a date.'''
    wb = xls.load_workbook(filename=file_path)
    ws = wb[sheet_name]

    line = top_line+1
    empty_lines_counter = 0
    stages_info = []

    while True:
        cell1 = ws.cell(row=line, column=planned_col)
        text1 = cell1.value

        cell2 = ws.cell(row=line, column=actual_col)
        text2 = cell2.value

        if text1 == None:
            empty_lines_counter += 1
            line += 1

        elif text1 != None:
            substrings_planned = text1.split(' ')       #['dd/mm', 'até', 'dd/mm']  planned dates
            begin_list_planned = substrings_planned[0].split('/')       #['dd', 'mm']
            finish_list_planned = substrings_planned[2].split('/')      #['dd', 'mm']
            dates_planned = [time.date(year, int(begin_list_planned[1]), int(begin_list_planned[0])), 
                        time.date(year, int(finish_list_planned[1]), int(finish_list_planned[0]))]

            if text2 != None:
                substrings_actual = text2.split(' ')        #same but for the executed dates
                begin_list_actual = substrings_actual[0].split('/')
                finish_list_actual = substrings_actual[2].split('/')
                dates_actual = [time.date(year, int(begin_list_actual[1]), int(begin_list_actual[0])), 
                                time.date(year, int(finish_list_actual[1]), int(finish_list_actual[0]))]
            elif text2 == None:
                dates_actual = [None, None]
            
            info = [line]+dates_planned+dates_actual        #[line, planned begin date, planned finish date, actual begin date, actual finish date]
            stages_info.append(info)

            empty_lines_counter = 0
            line +=1


        if empty_lines_counter >= 5: 
            break

    return stages_info


def get_weeks(list):
    '''Input: list of the form [first day, last day], both datetime.Date objects.'''

    begin_date = list[0]
    finish_date = list[1]
    day = time.timedelta(days=1)
    week = time.timedelta(days=7)

    for i in range(0, 8):
        if begin_date.isoweekday() == 1: break          # gets last monday before everything begins
        begin_date -= day

    for i in range(0,8):
        if finish_date.isoweekday() == 5: break         #gets first friday after everything ends
        finish_date += day


    mondays = []
    monday = begin_date
    while finish_date - monday > time.timedelta():
        mondays.append(monday)
        monday += week
    
    fridays = []
    friday = finish_date
    while friday - begin_date > time.timedelta():
        fridays.append(friday)
        friday -= week
    fridays = fridays[::-1]     #places the fridays in cronological order
    

    weeks = [[mondays[i], fridays[i]] for i in range(0, len(mondays))]      #[[monday1, friday1], [mon2, fri2], ...]
    return weeks

def get_first_n_last_day(info):
    '''Input: list from get_stage_dates, containing lists of the form [line, planned begin date, planned finish date, actual begin date, actual finish date]
    Returns a list containing two datetime.Date objects, representing the first and last day of construction.
    Main use is to get information for the get_weeks function, which needs these endpoints.'''
    first = info[0][1]
    last = info[-1][2]

    for i in range(0, len(info)):
        plan_begin = info[i][1]
        if plan_begin - first < time.timedelta():
            first = plan_begin
        
        plan_finish = info[i][2]
        actual_finish = info[i][4]
        if actual_finish == None:
            if plan_finish - last > time.timedelta():
                last = plan_finish
        else:
            if actual_finish - last > time.timedelta():
                last = actual_finish

    return [first, last]

def write_n_paint(file_path_main, sheet_name, top_line, first_colored_col, info, highlighted_lines):
    wb = xls.load_workbook(filename=file_path_main, read_only=False)
    ws = wb[sheet_name]

    first_last_days = get_first_n_last_day(info=info)
    weeks = get_weeks(list=first_last_days)

    planned_fill = PatternFill(start_color='99CCFF', patternType='solid')
    actual_fill = PatternFill(start_color='FF0000', patternType='solid')
    highlighted_fill = PatternFill(start_color='55FF55', patternType='solid')
    WHITE = PatternFill(start_color='FFFFFF', patternType='solid')

    week_time = time.timedelta(days=7)

    #write the weeks on the top line
    for i in range(0, len(weeks)):
        cell = ws.cell(row=top_line, column= first_colored_col + i)
        week = weeks[i]
        cell.value = f'{week[0].day}/{week[0].month}/{week[0].year - 2000} até {week[1].day}/{week[1].month}/{week[1 ].year - 2000}'    #print dates in "dd/mm/yy até dd/mm/yy format"


    #go through each line and paint
    for i in range(0, len(info)):
        line_info = info[i]
        line = line_info[0]
        begin_planned = line_info[1]
        finish_planned = line_info[2]
        begin_actual = line_info[3]
        finish_actual = line_info[4]

        #find begin_planned week
        for j in range(0, len(weeks)):
            if begin_planned - weeks[j][0] < week_time:
                break
        #find finish_planned week
        for k in range(0, len(weeks)):
            if finish_planned - weeks[k][0] < week_time:
                break  
        #fill planned weeks
        col_delta_set_planned = {x for x in range(j, k+1)}
        for col_delta in range(0, len(weeks)):
            cell = ws.cell(row=line, column=first_colored_col + col_delta)
            if col_delta in col_delta_set_planned:     #fill planned weeks
                if line in highlighted_lines:
                    cell.fill = highlighted_fill
                else:
                    cell.fill = planned_fill
            else:                               #paint previous schedule with white
                cell.fill = WHITE

        if begin_actual != None and finish_actual != None:
            #find begin_actual week
            for j in range(0, len(weeks)):
                if begin_actual - weeks[j][0] < week_time:
                    break
            #find finish_actual week
            for k in range(0, len(weeks)):
                if finish_actual - weeks[k][0] < week_time:
                    break
            #fill actual weeks
            col_delta_set_actual = {x for x in range(j, k+1)}-col_delta_set_planned
            if len(col_delta_set_actual) != 0:
                for col_delta in col_delta_set_actual:
                    cell = ws.cell(row=line, column=first_colored_col + col_delta)
                    cell.fill = actual_fill


    file_path = file_path_main.split('.')
    file_path.insert(-1, 'Copy')        #changes the name
    file_path[-1] = '.xlsx'             #puts back the . in front of xlsx
    file_path = ''.join(file_path)      #doesn't erase the original data, creates a copy file
    wb.save(filename=file_path)         #save edits to copy file



if __name__ == '__main__':
    try:
        print(User_message)

        file_path_main = input(path_message)
        sheet_name = input(sheet_name_message)

        planned_column = int(input(planned_column_message))
        actual_column = int(input(actual_column_message))
        first_colored_column = int(input(first_colored_column_message))

        top_line = int(input(top_line_message))
        highlighted_lines_input = input(highlighted_lines_message).split(',')
        if highlighted_lines_input != ['']:
            highlighted_lines = [int(x) for x in highlighted_lines_input]
        else:
            highlighted_lines = []

        info = get_stage_dates(file_path_main, sheet_name, planned_column, actual_column, top_line)
        write_n_paint(file_path_main, sheet_name, top_line, first_colored_column, info, highlighted_lines)

        print('\n\n', final_user_message)
        sleep(10)

    except Exception as exception:
        print(f'\nOcorreu um erro: {exception}')
        print('Tente novamente, ou, se o erro persistir, tire uma foto dessa mensagem de erro e mostre para um técnico que estiver acessível.')
        sleep(15)